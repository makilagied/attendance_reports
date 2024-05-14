import pandas as pd
from openpyxl import load_workbook, Workbook
import os
import getpass



class ExcelHandler:
    def save_to_excel(self, data_frame, file_path, sheet_name=None):
        successful_inserts = []
        unsuccessful_inserts = []

        try:
            # Try to load the existing workbook
            workbook = load_workbook(file_path)
        except FileNotFoundError:
            # If the file doesn't exist, create a new workbook
            workbook = Workbook()

        # Select the active sheet (first sheet by default)
        if sheet_name is None:
            sheet = workbook.active
        else:
            # Create a new sheet with the specified name or select an existing sheet
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.create_sheet(title=sheet_name)
                sheet.append(list(data_frame.columns))

        # Get existing data from the sheet
        existing_data = [list(row) for row in sheet.iter_rows(values_only=True)]

        #  # Remove the time component from 'Access Date' column
        # data_frame['Access Date'] = data_frame['Access Date'].dt.date

        # Check for duplicates and append new data at the bottom
        for _, row in data_frame.iterrows():
            row_list = row.values.tolist()

            if row_list not in existing_data:
                sheet.append(row_list)
                successful_inserts.append(row_list)
            else:
                unsuccessful_inserts.append(row_list)

        # Save the workbook
        workbook.save(file_path)

        # Print logs
        if len(successful_inserts) > 0:
            print(f"Successful Inserts: {len(successful_inserts)} out of {len(data_frame)}")
            print("DataFrame appended to Excel file.")
        else:
            print(
                f"\nUnsuccessful Inserts (Already in the Excel sheet): {len(unsuccessful_inserts)} out of {len(data_frame)}")
            for entry in unsuccessful_inserts:
                print(entry)
